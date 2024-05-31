import openpyxl
from openpyxl.styles.fills import PatternFill


def rows_Count(file, sheetKaNaam):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetKaNaam]
    rows = sheet.max_row
    return rows


def columns_Count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def write_data(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    # passing the data taken by user to a cell using "value" method
    sheet.cell(row=rowno, column=colno).value = data
    # whenever passing some data or styles from the code, always save the workbook
    workbook.save(file)


def read_data(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowno, column=colno).value


def fill_green_color(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    # storing the color details in a variable
    greenfill = PatternFill(start_color='60b212', end_color='60b212', patternType='solid')
    #step to add the color
    sheet.cell(row=rownum, column=colnum).fill = greenfill
    # whenever passing some data or styles from the code, always save the workbook
    workbook.save(file)


def fill_red_color(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    # storing the color details in a variable
    redfill = PatternFill(start_color='ff0000', end_color='ff0000', patternType='solid')
    #step to add the color
    sheet.cell(row=rownum, column=colnum).fill = redfill
    # whenever passing some data or styles from the code, always save the workbook
    workbook.save(file)
